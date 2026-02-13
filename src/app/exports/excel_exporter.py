# src/app/exports/excel_exporter.py
"""
Excel export with Bloomberg Finance Standard styling.

Creates a professional Excel workbook with 4 worksheets:
1. Trade Log - Detailed trade history
2. FIFO Cost Basis - Cost basis lots for user reference
3. Weekly Summary - Data for nobelltrading.com Saturday posting
4. Performance Dashboard - Monthly summaries and metrics
"""

from __future__ import annotations

import sqlite3
import os
import logging
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import List, Dict, Any, Optional, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

from app.constants import ExcelStyles, TradeThresholds, Defaults
from app.api.positions import get_schwab_positions
from app.db.cost_basis_db import get_avg_gain_for_sell
from app.cost_basis import extract_underlying

logger = logging.getLogger(__name__)


@dataclass
class TradeRecord:
    """Processed trade record for export."""
    symbol: str
    underlying: str
    asset_type: str
    action: str
    quantity: float
    filled_quantity: float
    position_remaining: int
    price: float
    status: str
    entry_date: str
    close_date: Optional[str]
    description: Optional[str]
    pl_value: float
    gain_pct: Optional[float]
    outcome: str  # WIN, LOSS, BREAKEVEN, UNKNOWN


@dataclass
class WeeklySummary:
    """Weekly performance summary for site posting."""
    week_ending: str
    total_trades: int
    wins: int
    losses: int
    breakevens: int
    win_rate: float
    total_pl: float
    best_trade_pct: Optional[float]
    worst_trade_pct: Optional[float]


class ExcelStyler:
    """Bloomberg Finance Standard styling for Excel."""

    def __init__(self):
        # Header style
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(
            start_color=ExcelStyles.HEADER_BG_COLOR,
            end_color=ExcelStyles.HEADER_BG_COLOR,
            fill_type="solid"
        )
        self.header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Borders
        self.thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        # Row alternating fills
        self.row_fill_1 = PatternFill(
            start_color=ExcelStyles.ROW_ALT_COLOR_1,
            end_color=ExcelStyles.ROW_ALT_COLOR_1,
            fill_type="solid"
        )
        self.row_fill_2 = PatternFill(
            start_color=ExcelStyles.ROW_ALT_COLOR_2,
            end_color=ExcelStyles.ROW_ALT_COLOR_2,
            fill_type="solid"
        )

        # Conditional formatting for outcomes
        self.win_fill = PatternFill(
            start_color=ExcelStyles.WIN_COLOR,
            end_color=ExcelStyles.WIN_COLOR,
            fill_type="solid"
        )
        self.win_font = Font(color=ExcelStyles.WIN_FONT, bold=True)

        self.loss_fill = PatternFill(
            start_color=ExcelStyles.LOSS_COLOR,
            end_color=ExcelStyles.LOSS_COLOR,
            fill_type="solid"
        )
        self.loss_font = Font(color=ExcelStyles.LOSS_FONT, bold=True)

        self.breakeven_fill = PatternFill(
            start_color=ExcelStyles.BREAKEVEN_COLOR,
            end_color=ExcelStyles.BREAKEVEN_COLOR,
            fill_type="solid"
        )
        self.breakeven_font = Font(color=ExcelStyles.BREAKEVEN_FONT, bold=True)

        # Total row style
        self.total_font = Font(bold=True, size=12)

    def apply_header_row(self, ws, headers: List[str], row: int = 1):
        """Apply Bloomberg-style header formatting."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_align
            cell.border = self.thin_border

    def apply_data_row(self, ws, row_num: int, num_cols: int, outcome: Optional[str] = None):
        """Apply alternating row colors and outcome highlighting."""
        # Alternating row fill
        base_fill = self.row_fill_1 if row_num % 2 == 0 else self.row_fill_2

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply base alternating color
            if outcome is None:
                cell.fill = base_fill

    def apply_outcome_style(self, cell, outcome: str):
        """Apply win/loss/breakeven styling to a cell."""
        if outcome == "WIN":
            cell.fill = self.win_fill
            cell.font = self.win_font
        elif outcome == "LOSS":
            cell.fill = self.loss_fill
            cell.font = self.loss_font
        elif outcome == "BREAKEVEN":
            cell.fill = self.breakeven_fill
            cell.font = self.breakeven_font

    def auto_size_columns(self, ws, headers: List[str], min_width: int = 10, max_width: int = 30):
        """Auto-size columns based on content."""
        for col in range(1, len(headers) + 1):
            max_length = len(headers[col - 1])
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            width = max(min_width, min(max_length + 2, max_width))
            ws.column_dimensions[get_column_letter(col)].width = width


class ExcelExporter:
    """
    Export trades to Bloomberg-style Excel workbook.

    Creates 4 worksheets:
    1. Trade Log - Detailed trade history with P/L
    2. FIFO Cost Basis - Cost basis lots and matches
    3. Weekly Summary - Data for nobelltrading.com
    4. Performance Dashboard - Monthly metrics
    """

    def __init__(self, db_path: str = None, output_dir: str = None):
        """
        Initialize exporter.

        Args:
            db_path: Path to SQLite database
            output_dir: Directory for output files
        """
        self.db_path = db_path or os.environ.get("DB_PATH", Defaults.DB_PATH)
        self.output_dir = output_dir or os.environ.get("OUTPUT_DIR", "/data")
        self.styler = ExcelStyler()

    def _get_connection(self) -> sqlite3.Connection:
        """Get database connection."""
        return sqlite3.connect(self.db_path)

    def _get_month_date_range(self, year: int, month: int) -> Tuple[str, str]:
        """Get start and end date strings for a given month."""
        start = f"{year}-{month:02d}-01"
        if month == 12:
            end = f"{year + 1}-01-01"
        else:
            end = f"{year}-{month + 1:02d}-01"
        return start, end

    def _classify_trade(self, gain_pct: Optional[float]) -> str:
        """Classify trade as WIN, LOSS, or BREAKEVEN using -10% to +10% rule."""
        return TradeThresholds.classify_trade(gain_pct)

    def _fetch_trades(self, conn: sqlite3.Connection, year: int, month: int) -> List[TradeRecord]:
        """Fetch and process trades for a month."""
        start_date, end_date = self._get_month_date_range(year, month)

        cursor = conn.execute("""
            SELECT
                order_id, symbol, asset_type, instruction, quantity,
                filled_quantity, remaining_quantity, price, status,
                entered_time, close_time, description
            FROM trades
            WHERE entered_time >= ? AND entered_time < ?
            ORDER BY entered_time ASC
        """, (start_date, end_date))
        rows = cursor.fetchall()

        # Get positions for remaining calculation (may fail without API credentials)
        try:
            _, positions_by_symbol = get_schwab_positions()
        except Exception as e:
            logger.warning(f"Could not fetch positions (API not available): {e}")
            positions_by_symbol = {}

        trades = []
        for row in rows:
            (order_id, symbol, asset_type, instruction, quantity,
             filled_qty, remaining_qty, price, status, entered, closed, desc) = row

            # Handle price - may be string "FILLED" in old data or numeric in new data
            try:
                price_val = float(price) if price and str(price).replace('.', '').isdigit() else 0
            except (ValueError, TypeError):
                price_val = 0

            # Calculate P/L value
            multiplier = 100 if asset_type == "OPTION" else 1
            filled = float(filled_qty) if filled_qty else float(quantity or 0)
            trade_value = price_val * filled * multiplier

            # Determine instruction - handle both old format (in asset_type) and new format
            instr = instruction or asset_type or ""
            if "SELL" in instr.upper():
                pl = trade_value
            elif "BUY" in instr.upper():
                pl = -trade_value
            else:
                pl = 0

            # Get position remaining
            underlying = extract_underlying(symbol)
            position_remaining = positions_by_symbol.get(underlying, 0)

            # Get gain percentage for sell orders
            gain_pct = None
            if "SELL" in instr.upper():
                gain_pct = get_avg_gain_for_sell(conn, order_id)

            # Classify outcome
            outcome = self._classify_trade(gain_pct) if gain_pct is not None else "UNKNOWN"

            trades.append(TradeRecord(
                symbol=symbol,
                underlying=underlying,
                asset_type=asset_type if asset_type in ("OPTION", "EQUITY") else "OPTION",
                action=instr,
                quantity=float(quantity or 0),
                filled_quantity=filled,
                position_remaining=position_remaining,
                price=price_val,
                status=status if status in ("FILLED", "CANCELED", "PENDING") else "FILLED",
                entry_date=entered,
                close_date=closed,
                description=desc,
                pl_value=pl,
                gain_pct=gain_pct,
                outcome=outcome
            ))

        return trades

    def _fetch_cost_basis_lots(self, conn: sqlite3.Connection, year: int, month: int) -> List[tuple]:
        """Fetch cost basis lots for a month."""
        try:
            start, end = self._get_month_date_range(year, month)
            cursor = conn.execute("""
                SELECT lot_id, order_id, symbol, underlying, quantity, remaining_qty,
                       avg_cost, entered_time, created_at
                FROM cost_basis_lots
                WHERE entered_time >= ? AND entered_time < ?
                ORDER BY entered_time DESC
            """, (start, end))
            return cursor.fetchall()
        except sqlite3.OperationalError as e:
            logger.warning(f"Could not fetch cost basis lots: {e}")
            return []

    def _fetch_lot_matches(self, conn: sqlite3.Connection, year: int, month: int) -> List[tuple]:
        """Fetch lot matches for a month."""
        try:
            start, end = self._get_month_date_range(year, month)
            cursor = conn.execute("""
                SELECT m.match_id, m.sell_order_id, m.lot_id, m.quantity,
                       m.cost_basis, m.sell_price, m.gain_pct, m.gain_amount, m.matched_at,
                       l.symbol, l.underlying
                FROM lot_matches m
                JOIN cost_basis_lots l ON m.lot_id = l.lot_id
                WHERE m.matched_at >= ? AND m.matched_at < ?
                ORDER BY m.matched_at DESC
            """, (start, end))
            return cursor.fetchall()
        except sqlite3.OperationalError as e:
            logger.warning(f"Could not fetch lot matches: {e}")
            return []

    def _calculate_weekly_summaries(self, trades: List[TradeRecord]) -> List[WeeklySummary]:
        """Calculate weekly summaries from trades."""
        # Group trades by week
        weeks: Dict[str, List[TradeRecord]] = {}
        for trade in trades:
            if not trade.entry_date:
                continue
            try:
                dt = datetime.fromisoformat(trade.entry_date.replace("Z", "+00:00"))
                # Get week ending (Saturday)
                days_until_saturday = (5 - dt.weekday()) % 7
                week_end = dt + timedelta(days=days_until_saturday)
                week_key = week_end.strftime("%Y-%m-%d")
                if week_key not in weeks:
                    weeks[week_key] = []
                weeks[week_key].append(trade)
            except (ValueError, AttributeError):
                continue

        summaries = []
        for week_ending, week_trades in sorted(weeks.items()):
            # Only count closing trades for win/loss stats
            closing_trades = [t for t in week_trades if "SELL" in t.action.upper() or "CLOSE" in t.action.upper()]

            wins = sum(1 for t in closing_trades if t.outcome == "WIN")
            losses = sum(1 for t in closing_trades if t.outcome == "LOSS")
            breakevens = sum(1 for t in closing_trades if t.outcome == "BREAKEVEN")

            total_closing = wins + losses + breakevens
            # Win rate excludes breakevens (or treats them as neutral)
            decisive_trades = wins + losses
            win_rate = (wins / decisive_trades * 100) if decisive_trades > 0 else 0

            total_pl = sum(t.pl_value for t in week_trades)

            # Best/worst trades
            gains = [t.gain_pct for t in closing_trades if t.gain_pct is not None]
            best = max(gains) if gains else None
            worst = min(gains) if gains else None

            summaries.append(WeeklySummary(
                week_ending=week_ending,
                total_trades=len(week_trades),
                wins=wins,
                losses=losses,
                breakevens=breakevens,
                win_rate=win_rate,
                total_pl=total_pl,
                best_trade_pct=best,
                worst_trade_pct=worst
            ))

        return summaries

    def _create_trade_log_sheet(self, wb: openpyxl.Workbook, trades: List[TradeRecord],
                                 month_name: str) -> None:
        """Create the Trade Log worksheet."""
        ws = wb.active
        ws.title = f"Trade Log"

        headers = [
            "Date/Time", "Symbol", "Underlying", "Action", "Qty", "Filled",
            "Price", "Total Value", "P/L ($)", "P/L %", "Outcome", "Status", "Notes"
        ]

        self.styler.apply_header_row(ws, headers)

        running_total = 0
        for row_idx, trade in enumerate(trades, 2):
            running_total += trade.pl_value

            # Apply base row styling FIRST (borders, alignment, alternating colors)
            self.styler.apply_data_row(ws, row_idx, len(headers))

            # Then populate data
            ws.cell(row=row_idx, column=1, value=trade.entry_date)
            ws.cell(row=row_idx, column=2, value=trade.symbol)
            ws.cell(row=row_idx, column=3, value=trade.underlying)
            ws.cell(row=row_idx, column=4, value=trade.action)
            ws.cell(row=row_idx, column=5, value=trade.quantity)
            ws.cell(row=row_idx, column=6, value=trade.filled_quantity)

            price_cell = ws.cell(row=row_idx, column=7, value=trade.price)
            price_cell.number_format = "$#,##0.00"

            multiplier = 100 if trade.asset_type == "OPTION" else 1
            total_value = trade.price * trade.filled_quantity * multiplier
            value_cell = ws.cell(row=row_idx, column=8, value=total_value)
            value_cell.number_format = "$#,##0.00"

            pl_cell = ws.cell(row=row_idx, column=9, value=trade.pl_value)
            pl_cell.number_format = "$#,##0.00"
            if trade.pl_value > 0:
                pl_cell.font = self.styler.win_font
            elif trade.pl_value < 0:
                pl_cell.font = self.styler.loss_font

            # Apply outcome styling AFTER base styling (overrides alternating color)
            if trade.gain_pct is not None:
                pct_cell = ws.cell(row=row_idx, column=10, value=trade.gain_pct / 100)
                pct_cell.number_format = "0.00%"
                self.styler.apply_outcome_style(pct_cell, trade.outcome)
            else:
                ws.cell(row=row_idx, column=10, value="N/A")

            outcome_cell = ws.cell(row=row_idx, column=11, value=trade.outcome)
            self.styler.apply_outcome_style(outcome_cell, trade.outcome)

            ws.cell(row=row_idx, column=12, value=trade.status)
            ws.cell(row=row_idx, column=13, value=trade.description or "")

        # Total row
        if trades:
            total_row = len(trades) + 2
            ws.cell(row=total_row, column=8, value="TOTAL P/L:").font = self.styler.total_font
            ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")

            total_pl = sum(t.pl_value for t in trades)
            total_cell = ws.cell(row=total_row, column=9, value=total_pl)
            total_cell.number_format = "$#,##0.00"
            total_cell.font = self.styler.total_font
            if total_pl >= 0:
                total_cell.fill = self.styler.win_fill
            else:
                total_cell.fill = self.styler.loss_fill

        self.styler.auto_size_columns(ws, headers)
        ws.freeze_panes = "A2"

    def _create_fifo_sheet(self, wb: openpyxl.Workbook, lots: List[tuple],
                           matches: List[tuple]) -> None:
        """Create the FIFO Cost Basis worksheet."""
        ws = wb.create_sheet("FIFO Cost Basis")

        # Section 1: Cost Basis Lots
        ws.cell(row=1, column=1, value="COST BASIS LOTS").font = Font(bold=True, size=14)
        ws.merge_cells("A1:I1")

        lot_headers = ["Lot ID", "Order ID", "Symbol", "Underlying", "Qty", "Remaining",
                       "Avg Cost", "Entry Time", "Created"]
        self.styler.apply_header_row(ws, lot_headers, row=2)

        for row_idx, lot in enumerate(lots, 3):
            for col_idx, value in enumerate(lot, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 7:  # Avg Cost
                    cell.number_format = "$#,##0.00"
                # Highlight remaining > 0
                if col_idx == 6 and value and value > 0:
                    cell.fill = self.styler.breakeven_fill
                    cell.font = Font(bold=True)
            self.styler.apply_data_row(ws, row_idx, len(lot_headers))

        # Section 2: FIFO Matches
        match_start_row = len(lots) + 5
        ws.cell(row=match_start_row, column=1, value="FIFO MATCHES (Sales)").font = Font(bold=True, size=14)
        ws.merge_cells(f"A{match_start_row}:K{match_start_row}")

        match_headers = ["Match ID", "Sell Order", "Lot ID", "Qty", "Cost Basis",
                         "Sell Price", "Gain %", "Gain $", "Matched At", "Symbol", "Underlying"]
        self.styler.apply_header_row(ws, match_headers, row=match_start_row + 1)

        total_gain = 0
        for row_idx, match in enumerate(matches, match_start_row + 2):
            for col_idx, value in enumerate(match, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                if col_idx in (5, 6):  # Cost Basis, Sell Price
                    cell.number_format = "$#,##0.00"
                if col_idx == 7 and value:  # Gain %
                    cell.number_format = "0.00%"
                    cell.value = value / 100
                    outcome = TradeThresholds.classify_trade(value)
                    self.styler.apply_outcome_style(cell, outcome)
                if col_idx == 8:  # Gain $
                    cell.number_format = "$#,##0.00"
                    if value and value > 0:
                        cell.font = self.styler.win_font
                    elif value and value < 0:
                        cell.font = self.styler.loss_font

            if match[7]:  # gain_amount
                total_gain += match[7]

            self.styler.apply_data_row(ws, row_idx, len(match_headers))

        # Total gain row
        if matches:
            total_row = match_start_row + 2 + len(matches)
            ws.cell(row=total_row, column=7, value="TOTAL GAIN:").font = self.styler.total_font
            ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
            gain_cell = ws.cell(row=total_row, column=8, value=total_gain)
            gain_cell.number_format = "$#,##0.00"
            gain_cell.font = self.styler.total_font
            if total_gain >= 0:
                gain_cell.fill = self.styler.win_fill
            else:
                gain_cell.fill = self.styler.loss_fill

        # Auto-size
        for col in range(1, max(len(lot_headers), len(match_headers)) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.freeze_panes = "A3"

    def _create_weekly_summary_sheet(self, wb: openpyxl.Workbook,
                                     summaries: List[WeeklySummary]) -> None:
        """Create the Weekly Summary worksheet for nobelltrading.com posting."""
        ws = wb.create_sheet("Weekly Summary")

        # Title for site
        ws.cell(row=1, column=1, value="NOBELL TRADING - WEEKLY PERFORMANCE").font = Font(bold=True, size=16)
        ws.merge_cells("A1:I1")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        headers = [
            "Week Ending", "Total Trades", "Wins", "Losses", "Break-even",
            "Win Rate %", "Total P/L", "Best Trade %", "Worst Trade %"
        ]
        self.styler.apply_header_row(ws, headers, row=3)

        for row_idx, summary in enumerate(summaries, 4):
            ws.cell(row=row_idx, column=1, value=summary.week_ending)
            ws.cell(row=row_idx, column=2, value=summary.total_trades)

            wins_cell = ws.cell(row=row_idx, column=3, value=summary.wins)
            wins_cell.font = self.styler.win_font

            losses_cell = ws.cell(row=row_idx, column=4, value=summary.losses)
            losses_cell.font = self.styler.loss_font

            be_cell = ws.cell(row=row_idx, column=5, value=summary.breakevens)
            be_cell.font = self.styler.breakeven_font

            wr_cell = ws.cell(row=row_idx, column=6, value=summary.win_rate / 100)
            wr_cell.number_format = "0.0%"
            if summary.win_rate >= 50:
                wr_cell.font = self.styler.win_font
            else:
                wr_cell.font = self.styler.loss_font

            pl_cell = ws.cell(row=row_idx, column=7, value=summary.total_pl)
            pl_cell.number_format = "$#,##0.00"
            if summary.total_pl >= 0:
                pl_cell.font = self.styler.win_font
            else:
                pl_cell.font = self.styler.loss_font

            if summary.best_trade_pct is not None:
                best_cell = ws.cell(row=row_idx, column=8, value=summary.best_trade_pct / 100)
                best_cell.number_format = "+0.0%;-0.0%"
                best_cell.font = self.styler.win_font
            else:
                ws.cell(row=row_idx, column=8, value="N/A")

            if summary.worst_trade_pct is not None:
                worst_cell = ws.cell(row=row_idx, column=9, value=summary.worst_trade_pct / 100)
                worst_cell.number_format = "+0.0%;-0.0%"
                worst_cell.font = self.styler.loss_font
            else:
                ws.cell(row=row_idx, column=9, value="N/A")

            self.styler.apply_data_row(ws, row_idx, len(headers))

        # Summary totals
        if summaries:
            total_row = len(summaries) + 5
            ws.cell(row=total_row, column=1, value="MONTH TOTAL").font = self.styler.total_font

            total_trades = sum(s.total_trades for s in summaries)
            total_wins = sum(s.wins for s in summaries)
            total_losses = sum(s.losses for s in summaries)
            total_be = sum(s.breakevens for s in summaries)
            total_pl = sum(s.total_pl for s in summaries)

            decisive = total_wins + total_losses
            overall_wr = (total_wins / decisive * 100) if decisive > 0 else 0

            ws.cell(row=total_row, column=2, value=total_trades).font = self.styler.total_font
            ws.cell(row=total_row, column=3, value=total_wins).font = self.styler.total_font
            ws.cell(row=total_row, column=4, value=total_losses).font = self.styler.total_font
            ws.cell(row=total_row, column=5, value=total_be).font = self.styler.total_font

            wr_total = ws.cell(row=total_row, column=6, value=overall_wr / 100)
            wr_total.number_format = "0.0%"
            wr_total.font = self.styler.total_font

            pl_total = ws.cell(row=total_row, column=7, value=total_pl)
            pl_total.number_format = "$#,##0.00"
            pl_total.font = self.styler.total_font
            if total_pl >= 0:
                pl_total.fill = self.styler.win_fill
            else:
                pl_total.fill = self.styler.loss_fill

        # Add note about break-even rule
        note_row = len(summaries) + 7
        ws.cell(row=note_row, column=1,
                value="Note: Break-even = trades with gain between -10% and +10%").font = Font(italic=True)
        ws.merge_cells(f"A{note_row}:I{note_row}")

        self.styler.auto_size_columns(ws, headers)
        ws.freeze_panes = "A4"

    def _create_performance_dashboard(self, wb: openpyxl.Workbook,
                                       trades: List[TradeRecord],
                                       summaries: List[WeeklySummary],
                                       month_name: str) -> None:
        """Create the Performance Dashboard worksheet."""
        ws = wb.create_sheet("Performance Dashboard")

        # Title
        ws.cell(row=1, column=1, value=f"PERFORMANCE DASHBOARD - {month_name}").font = Font(bold=True, size=16)
        ws.merge_cells("A1:E1")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        # Key metrics section
        ws.cell(row=3, column=1, value="KEY METRICS").font = Font(bold=True, size=14)

        closing_trades = [t for t in trades if "SELL" in t.action.upper() or "CLOSE" in t.action.upper()]

        metrics = [
            ("Total Trades", len(trades)),
            ("Closing Trades", len(closing_trades)),
            ("Wins (>+10%)", sum(1 for t in closing_trades if t.outcome == "WIN")),
            ("Losses (<-10%)", sum(1 for t in closing_trades if t.outcome == "LOSS")),
            ("Break-even (-10% to +10%)", sum(1 for t in closing_trades if t.outcome == "BREAKEVEN")),
        ]

        wins = sum(1 for t in closing_trades if t.outcome == "WIN")
        losses = sum(1 for t in closing_trades if t.outcome == "LOSS")
        decisive = wins + losses
        win_rate = (wins / decisive * 100) if decisive > 0 else 0
        metrics.append(("Win Rate", f"{win_rate:.1f}%"))

        total_pl = sum(t.pl_value for t in trades)
        metrics.append(("Total P/L", f"${total_pl:,.2f}"))

        gains = [t.gain_pct for t in closing_trades if t.gain_pct is not None]
        if gains:
            avg_gain = sum(gains) / len(gains)
            metrics.append(("Average Gain %", f"{avg_gain:.2f}%"))
            metrics.append(("Best Trade", f"{max(gains):.2f}%"))
            metrics.append(("Worst Trade", f"{min(gains):.2f}%"))

        for row_idx, (label, value) in enumerate(metrics, 4):
            label_cell = ws.cell(row=row_idx, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="right")

            value_cell = ws.cell(row=row_idx, column=2, value=value)
            if "Win Rate" in label or "P/L" in label or "Gain" in label or "Trade" in label:
                if isinstance(value, str) and ("-" in value or "Loss" in label.lower()):
                    value_cell.font = self.styler.loss_font
                elif isinstance(value, str) and ("Win" in label or "+" in value):
                    value_cell.font = self.styler.win_font
                else:
                    value_cell.font = Font(bold=True)

        # Weekly breakdown
        weekly_start = len(metrics) + 6
        ws.cell(row=weekly_start, column=1, value="WEEKLY BREAKDOWN").font = Font(bold=True, size=14)

        ws_headers = ["Week", "Trades", "Win Rate", "P/L"]
        for col, header in enumerate(ws_headers, 1):
            cell = ws.cell(row=weekly_start + 1, column=col, value=header)
            cell.font = self.styler.header_font
            cell.fill = self.styler.header_fill

        for row_idx, summary in enumerate(summaries, weekly_start + 2):
            ws.cell(row=row_idx, column=1, value=summary.week_ending)
            ws.cell(row=row_idx, column=2, value=summary.total_trades)

            wr_cell = ws.cell(row=row_idx, column=3, value=summary.win_rate / 100)
            wr_cell.number_format = "0.0%"

            pl_cell = ws.cell(row=row_idx, column=4, value=summary.total_pl)
            pl_cell.number_format = "$#,##0.00"

        # Auto-size
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def export(self, year: int = None, month: int = None) -> str:
        """
        Export trades to Excel workbook.

        Args:
            year: Year to export (defaults to current)
            month: Month to export (defaults to current)

        Returns:
            Path to created Excel file
        """
        # Default to current month
        if year is None or month is None:
            now = datetime.now()
            year = now.year
            month = now.month

        month_name = datetime(year, month, 1).strftime("%B %Y")
        filename = f"trades_{year}-{month:02d}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        conn = self._get_connection()
        try:
            # Fetch data
            trades = self._fetch_trades(conn, year, month)
            lots = self._fetch_cost_basis_lots(conn, year, month)
            matches = self._fetch_lot_matches(conn, year, month)
            summaries = self._calculate_weekly_summaries(trades)

            # Create workbook
            wb = openpyxl.Workbook()

            # Create sheets
            self._create_trade_log_sheet(wb, trades, month_name)
            self._create_fifo_sheet(wb, lots, matches)
            self._create_weekly_summary_sheet(wb, summaries)
            self._create_performance_dashboard(wb, trades, summaries, month_name)

            # Save
            wb.save(filepath)

            # Log summary
            total_pl = sum(t.pl_value for t in trades)
            closing_trades = [t for t in trades if "SELL" in t.action.upper() or "CLOSE" in t.action.upper()]
            wins = sum(1 for t in closing_trades if t.outcome == "WIN")
            losses = sum(1 for t in closing_trades if t.outcome == "LOSS")
            breakevens = sum(1 for t in closing_trades if t.outcome == "BREAKEVEN")

            logger.info(f"Excel export completed: {filepath}")
            logger.info(f"  Month: {month_name}")
            logger.info(f"  Trades: {len(trades)} | Closing: {len(closing_trades)}")
            logger.info(f"  Wins: {wins} | Losses: {losses} | Break-even: {breakevens}")
            logger.info(f"  Total P/L: ${total_pl:,.2f}")

            return filepath

        finally:
            conn.close()

    def _fetch_all_trades(self, conn: sqlite3.Connection) -> List[TradeRecord]:
        """Fetch and process ALL trades without date filtering."""
        cursor = conn.execute("""
            SELECT
                order_id, symbol, asset_type, instruction, quantity,
                filled_quantity, remaining_quantity, price, status,
                entered_time, close_time, description
            FROM trades
            ORDER BY entered_time ASC
        """)
        rows = cursor.fetchall()

        # Get positions for remaining calculation (may fail without API credentials)
        try:
            _, positions_by_symbol = get_schwab_positions()
        except Exception as e:
            logger.warning(f"Could not fetch positions (API not available): {e}")
            positions_by_symbol = {}

        trades = []
        for row in rows:
            (order_id, symbol, asset_type, instruction, quantity,
             filled_qty, remaining_qty, price, status, entered, closed, desc) = row

            # Handle price - may be string "FILLED" in old data or numeric in new data
            try:
                price_val = float(price) if price and str(price).replace('.', '').isdigit() else 0
            except (ValueError, TypeError):
                price_val = 0

            # Calculate P/L value
            multiplier = 100 if asset_type == "OPTION" else 1
            filled = float(filled_qty) if filled_qty else float(quantity or 0)
            trade_value = price_val * filled * multiplier

            # Determine instruction - handle both old format (in asset_type) and new format
            instr = instruction or asset_type or ""
            if "SELL" in instr.upper():
                pl = trade_value
            elif "BUY" in instr.upper():
                pl = -trade_value
            else:
                pl = 0

            # Get position remaining
            underlying = extract_underlying(symbol)
            position_remaining = positions_by_symbol.get(underlying, 0)

            # Get gain percentage for sell orders
            gain_pct = None
            if "SELL" in instr.upper():
                gain_pct = get_avg_gain_for_sell(conn, order_id)

            # Classify outcome
            outcome = self._classify_trade(gain_pct) if gain_pct is not None else "UNKNOWN"

            trades.append(TradeRecord(
                symbol=symbol,
                underlying=underlying,
                asset_type=asset_type if asset_type in ("OPTION", "EQUITY") else "OPTION",
                action=instr,
                quantity=float(quantity or 0),
                filled_quantity=filled,
                position_remaining=position_remaining,
                price=price_val,
                status=status if status in ("FILLED", "CANCELED", "PENDING") else "FILLED",
                entry_date=entered,
                close_date=closed,
                description=desc,
                pl_value=pl,
                gain_pct=gain_pct,
                outcome=outcome
            ))

        return trades

    def _fetch_all_cost_basis_lots(self, conn: sqlite3.Connection) -> List[tuple]:
        """Fetch ALL cost basis lots without date filtering."""
        try:
            cursor = conn.execute("""
                SELECT lot_id, order_id, symbol, underlying, quantity, remaining_qty,
                       avg_cost, entered_time, created_at
                FROM cost_basis_lots
                ORDER BY entered_time DESC
            """)
            return cursor.fetchall()
        except sqlite3.OperationalError as e:
            logger.warning(f"Could not fetch cost basis lots: {e}")
            return []

    def _fetch_all_lot_matches(self, conn: sqlite3.Connection) -> List[tuple]:
        """Fetch ALL lot matches without date filtering."""
        try:
            cursor = conn.execute("""
                SELECT m.match_id, m.sell_order_id, m.lot_id, m.quantity,
                       m.cost_basis, m.sell_price, m.gain_pct, m.gain_amount, m.matched_at,
                       l.symbol, l.underlying
                FROM lot_matches m
                JOIN cost_basis_lots l ON m.lot_id = l.lot_id
                ORDER BY m.matched_at DESC
            """)
            return cursor.fetchall()
        except sqlite3.OperationalError as e:
            logger.warning(f"Could not fetch lot matches: {e}")
            return []

    def export_all(self) -> str:
        """
        Export ALL trades to a single Excel workbook with full styling.

        Returns:
            Path to created Excel file
        """
        filename = "trades_all.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        conn = self._get_connection()
        try:
            # Fetch all data
            trades = self._fetch_all_trades(conn)
            lots = self._fetch_all_cost_basis_lots(conn)
            matches = self._fetch_all_lot_matches(conn)
            summaries = self._calculate_weekly_summaries(trades)

            # Get date range
            if trades:
                dates = [t.entry_date[:10] for t in trades if t.entry_date]
                date_range = f"{min(dates)} to {max(dates)}" if dates else "All Time"
            else:
                date_range = "All Time"

            # Create workbook
            wb = openpyxl.Workbook()

            # Create sheets
            self._create_trade_log_sheet(wb, trades, f"All Trades ({date_range})")
            self._create_fifo_sheet(wb, lots, matches)
            self._create_weekly_summary_sheet(wb, summaries)
            self._create_performance_dashboard(wb, trades, summaries, "All Time")

            # Save
            wb.save(filepath)

            # Log summary
            total_pl = sum(t.pl_value for t in trades)
            closing_trades = [t for t in trades if "SELL" in t.action.upper() or "CLOSE" in t.action.upper()]
            wins = sum(1 for t in closing_trades if t.outcome == "WIN")
            losses = sum(1 for t in closing_trades if t.outcome == "LOSS")
            breakevens = sum(1 for t in closing_trades if t.outcome == "BREAKEVEN")

            logger.info(f"Excel export completed: {filepath}")
            logger.info(f"  Date Range: {date_range}")
            logger.info(f"  Trades: {len(trades)} | Closing: {len(closing_trades)}")
            logger.info(f"  Wins: {wins} | Losses: {losses} | Break-even: {breakevens}")
            logger.info(f"  Total P/L: ${total_pl:,.2f}")

            return filepath

        finally:
            conn.close()


def export_trades(year: int = None, month: int = None) -> str:
    """
    Convenience function to export trades.

    Args:
        year: Year to export (defaults to current)
        month: Month to export (defaults to current)

    Returns:
        Path to created Excel file
    """
    exporter = ExcelExporter()
    return exporter.export(year, month)


def export_all_trades(output_dir: str = None, db_path: str = None) -> str:
    """
    Export all trades to a single Excel file with full styling.

    Args:
        output_dir: Output directory for the file
        db_path: Path to the database

    Returns:
        Path to created Excel file
    """
    exporter = ExcelExporter(db_path=db_path, output_dir=output_dir)
    return exporter.export_all()


if __name__ == "__main__":
    import sys

    if len(sys.argv) >= 3:
        year = int(sys.argv[1])
        month = int(sys.argv[2])
        export_trades(year, month)
    else:
        export_trades()
