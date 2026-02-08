#!/usr/bin/env python3
"""Export trades from SQLite to Excel file in data folder.

Creates monthly spreadsheets named trades_YYYY-MM.xlsx to keep files manageable.
Each month gets its own file. New trades are APPENDED to existing files (newest at bottom).
"""

import sqlite3
import os
import logging
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

from app.api.positions import get_schwab_positions
from app.db.cost_basis_db import get_avg_gain_for_sell
from app.cost_basis import extract_underlying

DB_PATH = os.environ.get("DB_PATH", "/data/trades.db")
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "/data")


def get_current_month_filter():
    """Get the current year and month for filtering."""
    now = datetime.now()
    return now.year, now.month


def get_month_date_range(year: int, month: int):
    """Get start and end date strings for a given month."""
    start = f"{year}-{month:02d}-01"
    # Handle month rollover for end date
    if month == 12:
        end = f"{year + 1}-01-01"
    else:
        end = f"{year}-{month + 1:02d}-01"
    return start, end


def get_cost_basis_lots(conn, year: int = None, month: int = None):
    """Get cost basis lots from database, optionally filtered by month."""
    try:
        if year and month:
            start, end = get_month_date_range(year, month)
            cursor = conn.execute("""
                SELECT lot_id, order_id, symbol, underlying, quantity, remaining_qty,
                       avg_cost, entered_time, created_at
                FROM cost_basis_lots
                WHERE entered_time >= ? AND entered_time < ?
                ORDER BY entered_time DESC
            """, (start, end))
        else:
            cursor = conn.execute("""
                SELECT lot_id, order_id, symbol, underlying, quantity, remaining_qty,
                       avg_cost, entered_time, created_at
                FROM cost_basis_lots
                ORDER BY entered_time DESC
            """)
        return cursor.fetchall()
    except sqlite3.OperationalError as e:
        logger.warning(f"Could not fetch cost basis lots (table may not exist): {e}")
        return []


def get_lot_matches(conn, year: int = None, month: int = None):
    """Get lot matches from database, optionally filtered by month."""
    try:
        if year and month:
            start, end = get_month_date_range(year, month)
            cursor = conn.execute("""
                SELECT m.match_id, m.sell_order_id, m.lot_id, m.quantity,
                       m.cost_basis, m.sell_price, m.gain_pct, m.gain_amount, m.matched_at,
                       l.symbol, l.underlying
                FROM lot_matches m
                JOIN cost_basis_lots l ON m.lot_id = l.lot_id
                WHERE m.matched_at >= ? AND m.matched_at < ?
                ORDER BY m.matched_at DESC
            """, (start, end))
        else:
            cursor = conn.execute("""
                SELECT m.match_id, m.sell_order_id, m.lot_id, m.quantity,
                       m.cost_basis, m.sell_price, m.gain_pct, m.gain_amount, m.matched_at,
                       l.symbol, l.underlying
                FROM lot_matches m
                JOIN cost_basis_lots l ON m.lot_id = l.lot_id
                ORDER BY m.matched_at DESC
            """)
        return cursor.fetchall()
    except sqlite3.OperationalError as e:
        logger.warning(f"Could not fetch lot matches (table may not exist): {e}")
        return []


def generate_filename(year: int, month: int) -> str:
    """Generate filename with month and year: trades_YYYY-MM.xlsx"""
    return f"trades_{year}-{month:02d}.xlsx"


def get_existing_entry_dates(ws):
    """Get set of entry dates already in the trades sheet (column 9 = Entry Date)."""
    existing = set()
    for row in range(2, ws.max_row + 1):
        entry_date = ws.cell(row=row, column=9).value
        if entry_date:
            existing.add(str(entry_date))
    return existing


def find_total_row(ws):
    """Find the row with 'TOTAL P/L:' label, or return None if not found."""
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=11).value == "TOTAL P/L:":
            return row
    return None


def export_trades(year: int = None, month: int = None):
    """Export trades to Excel. Appends new trades to existing file (newest at bottom)."""
    # Default to current month
    if year is None or month is None:
        year, month = get_current_month_filter()

    start_date, end_date = get_month_date_range(year, month)
    month_name = datetime(year, month, 1).strftime("%B %Y")
    filename = generate_filename(year, month)
    filepath = os.path.join(OUTPUT_DIR, filename)

    conn = sqlite3.connect(DB_PATH)

    # Get trades for the specified month - ORDER BY ASC so newest is last
    cursor = conn.execute("""
        SELECT
            order_id,
            symbol,
            asset_type,
            instruction,
            quantity,
            filled_quantity,
            remaining_quantity,
            price,
            status,
            entered_time,
            close_time,
            description
        FROM trades
        WHERE entered_time >= ? AND entered_time < ?
        ORDER BY entered_time ASC
    """, (start_date, end_date))
    rows = cursor.fetchall()

    # Get actual positions from Schwab (always current)
    schwab_positions, positions_by_symbol = get_schwab_positions()

    # Get cost basis data for this month
    cost_basis_lots = get_cost_basis_lots(conn, year, month)
    lot_matches = get_lot_matches(conn, year, month)

    # Calculate P/L per trade and add position remaining
    all_trades = []
    for row in rows:
        order_id, symbol, asset_type, instruction, quantity, filled_qty, remaining_qty, price, status, entered, closed, desc = row

        multiplier = 100 if asset_type == "OPTION" else 1
        filled = filled_qty if filled_qty else quantity
        trade_value = (price or 0) * filled * multiplier

        if instruction and "SELL" in instruction.upper():
            pl = trade_value
        elif instruction and "BUY" in instruction.upper():
            pl = -trade_value
        else:
            pl = 0

        # Get actual position remaining from Schwab (by underlying symbol)
        underlying = extract_underlying(symbol)
        position_remaining = positions_by_symbol.get(underlying, 0)

        # Get gain percentage for sell orders
        gain_pct = None
        if instruction and "SELL" in instruction.upper():
            gain_pct = get_avg_gain_for_sell(conn, order_id)

        all_trades.append((symbol, asset_type, instruction, quantity, filled_qty,
                           position_remaining, price, status, entered, closed, desc, pl, gain_pct))

    conn.close()

    # Check if file exists - if so, load and append
    file_exists = os.path.exists(filepath)
    new_trades_count = 0

    if file_exists:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active

        # Always remove the totals row if it exists (we'll re-add it at the end)
        total_row = find_total_row(ws)
        if total_row:
            ws.delete_rows(total_row)

        # Get existing entry dates to avoid duplicates
        existing_dates = get_existing_entry_dates(ws)

        # Filter to only new trades
        new_trades = [t for t in all_trades if str(t[8]) not in existing_dates]
        new_trades_count = len(new_trades)
        trades_with_pl = new_trades
    else:
        # Create new workbook
        wb = openpyxl.Workbook()
        trades_with_pl = all_trades
        new_trades_count = len(all_trades)

    # Style settings
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    profit_font = Font(color="006400", bold=True)
    loss_font = Font(color="8B0000", bold=True)

    headers = [
        "Symbol", "Asset Type", "Action", "Quantity", "Filled", "Position Left",
        "Price", "Status", "Entry Date", "Close Date", "Notes", "P/L ($)", "Gain %"
    ]

    # ===== SHEET 1: Monthly Trades =====
    ws = wb.active

    if not file_exists:
        # New file - set up headers
        ws.title = f"Trades {month_name}"
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        start_row = 2
    else:
        # Existing file - find where to append
        start_row = ws.max_row + 1 if trades_with_pl else ws.max_row

    # Add new trades
    for idx, row in enumerate(trades_with_pl):
        row_idx = start_row + idx
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Position Left column - highlight if > 0
            if col_idx == 6 and value and value > 0:
                cell.font = Font(bold=True, color="1F4E79")

            # P/L column
            if col_idx == 12:
                cell.number_format = "$#,##0.00"
                if value and value > 0:
                    cell.font = profit_font
                elif value and value < 0:
                    cell.font = loss_font

            # Gain % column
            if col_idx == 13 and value is not None:
                cell.number_format = "0.00%"
                cell.value = value / 100 if value else 0  # Convert to decimal for %
                if value and value > 0:
                    cell.font = profit_font
                elif value and value < 0:
                    cell.font = loss_font

        action = row[2] if row[2] else ""
        if "BUY" in action.upper():
            for col in range(1, len(headers) + 1):
                if col not in (6, 12, 13):
                    ws.cell(row=row_idx, column=col).fill = buy_fill
        elif "SELL" in action.upper():
            for col in range(1, len(headers) + 1):
                if col not in (6, 12, 13):
                    ws.cell(row=row_idx, column=col).fill = sell_fill

    # Calculate total P/L from ALL rows in sheet (not just new ones)
    total_pl = 0
    for row in range(2, ws.max_row + 1):
        pl_val = ws.cell(row=row, column=12).value
        if pl_val and isinstance(pl_val, (int, float)):
            total_pl += pl_val

    # Add totals row at the end
    if ws.max_row > 1:
        total_row_num = ws.max_row + 1
        ws.cell(row=total_row_num, column=11, value="TOTAL P/L:").font = Font(bold=True)
        ws.cell(row=total_row_num, column=11).alignment = Alignment(horizontal="right")
        total_cell = ws.cell(row=total_row_num, column=12, value=total_pl)
        total_cell.number_format = "$#,##0.00"
        total_cell.font = Font(bold=True, color="006400" if total_pl >= 0 else "8B0000")
        total_cell.border = thin_border

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col-1])
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)

    ws.freeze_panes = "A2"

    # ===== SHEET 2: Open Positions (from Schwab API - always current) =====
    # This sheet is always replaced with current data
    if "Open Positions" in wb.sheetnames:
        del wb["Open Positions"]
    ws2 = wb.create_sheet("Open Positions")

    pos_headers = ["Symbol", "Asset Type", "Quantity", "Avg Price", "Market Value", "Status"]

    for col, header in enumerate(pos_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_idx = 2
    total_qty = 0
    total_value = 0

    for pos in schwab_positions:
        ws2.cell(row=row_idx, column=1, value=pos["symbol"]).border = thin_border
        ws2.cell(row=row_idx, column=2, value=pos["asset_type"]).border = thin_border

        qty_cell = ws2.cell(row=row_idx, column=3, value=pos["quantity"])
        qty_cell.border = thin_border
        qty_cell.font = Font(bold=True)

        price_cell = ws2.cell(row=row_idx, column=4, value=pos["avg_price"])
        price_cell.border = thin_border
        price_cell.number_format = "$#,##0.00"

        val_cell = ws2.cell(row=row_idx, column=5, value=pos["market_value"])
        val_cell.border = thin_border
        val_cell.number_format = "$#,##0.00"

        status_cell = ws2.cell(row=row_idx, column=6, value="OPEN - TO SELL")
        status_cell.border = thin_border
        status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        status_cell.font = Font(bold=True)

        for col in range(1, 6):
            ws2.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

        total_qty += pos["quantity"]
        total_value += pos["market_value"]
        row_idx += 1

    # Total row
    if schwab_positions:
        ws2.cell(row=row_idx + 1, column=2, value="TOTAL:").font = Font(bold=True)
        ws2.cell(row=row_idx + 1, column=2).alignment = Alignment(horizontal="right")

        ws2.cell(row=row_idx + 1, column=3, value=total_qty).font = Font(bold=True, color="1F4E79")
        ws2.cell(row=row_idx + 1, column=3).border = thin_border

        total_val_cell = ws2.cell(row=row_idx + 1, column=5, value=total_value)
        total_val_cell.font = Font(bold=True, color="1F4E79")
        total_val_cell.border = thin_border
        total_val_cell.number_format = "$#,##0.00"

    for col in range(1, len(pos_headers) + 1):
        max_length = len(pos_headers[col-1])
        for row in range(2, row_idx):
            cell_value = ws2.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws2.column_dimensions[get_column_letter(col)].width = min(max_length + 3, 25)

    ws2.freeze_panes = "A2"

    # ===== SHEET 3: Cost Basis Lots (this month) =====
    # This sheet is always replaced with current data
    if "Cost Basis Lots" in wb.sheetnames:
        del wb["Cost Basis Lots"]
    ws3 = wb.create_sheet("Cost Basis Lots")

    lot_headers = ["Lot ID", "Order ID", "Symbol", "Underlying", "Quantity", "Remaining",
                   "Avg Cost", "Entry Time", "Created At"]

    for col, header in enumerate(lot_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, lot in enumerate(cost_basis_lots, 2):
        for col_idx, value in enumerate(lot, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Avg Cost column
            if col_idx == 7:
                cell.number_format = "$#,##0.00"

            # Remaining qty - highlight if > 0
            if col_idx == 6 and value and value > 0:
                cell.font = Font(bold=True, color="1F4E79")
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col in range(1, len(lot_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 15

    ws3.freeze_panes = "A2"

    # ===== SHEET 4: Lot Matches (FIFO Sales - this month) =====
    # This sheet is always replaced with current data
    if "FIFO Matches" in wb.sheetnames:
        del wb["FIFO Matches"]
    ws4 = wb.create_sheet("FIFO Matches")

    match_headers = ["Match ID", "Sell Order", "Lot ID", "Quantity", "Cost Basis",
                     "Sell Price", "Gain %", "Gain $", "Matched At", "Symbol", "Underlying"]

    for col, header in enumerate(match_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    total_gain = 0
    for row_idx, match in enumerate(lot_matches, 2):
        for col_idx, value in enumerate(match, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Cost Basis and Sell Price columns
            if col_idx in (5, 6):
                cell.number_format = "$#,##0.00"

            # Gain % column
            if col_idx == 7:
                cell.number_format = "0.00%"
                cell.value = value / 100 if value else 0
                if value and value > 0:
                    cell.font = profit_font
                elif value and value < 0:
                    cell.font = loss_font

            # Gain $ column
            if col_idx == 8:
                cell.number_format = "$#,##0.00"
                if value and value > 0:
                    cell.font = profit_font
                elif value and value < 0:
                    cell.font = loss_font

        if match[7]:  # gain_amount
            total_gain += match[7]

    # Total row for matches
    if lot_matches:
        total_row = len(lot_matches) + 2
        ws4.cell(row=total_row, column=7, value="TOTAL GAIN:").font = Font(bold=True)
        ws4.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")
        gain_cell = ws4.cell(row=total_row, column=8, value=total_gain)
        gain_cell.number_format = "$#,##0.00"
        gain_cell.font = Font(bold=True, color="006400" if total_gain >= 0 else "8B0000")
        gain_cell.border = thin_border

    for col in range(1, len(match_headers) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 15

    ws4.freeze_panes = "A2"

    # Save file
    wb.save(filepath)

    # Count total trades in sheet (excluding header and total rows)
    total_trades_in_file = ws.max_row - 2 if ws.max_row > 2 else ws.max_row - 1

    print(f"=" * 60)
    print(f"EXPORT: {month_name}")
    print(f"=" * 60)
    print(f"File: {filepath}")
    if file_exists:
        print(f"Mode: APPEND ({new_trades_count} new trades added)")
    else:
        print(f"Mode: NEW FILE CREATED")
    print(f"Total trades in file: {total_trades_in_file}")
    print(f"Monthly P/L: ${total_pl:,.2f}")
    print(f"Open positions: {len(schwab_positions)} ({total_qty:.0f} contracts)")
    print(f"Total market value: ${total_value:,.2f}")
    print(f"Cost basis lots (this month): {len(cost_basis_lots)}")
    print(f"FIFO matches (this month): {len(lot_matches)}")
    if lot_matches:
        print(f"Monthly realized gain: ${total_gain:,.2f}")
    print(f"=" * 60)

    return filepath


if __name__ == "__main__":
    import sys

    # Allow passing year and month as arguments: python export_trades.py 2026 2
    if len(sys.argv) >= 3:
        year = int(sys.argv[1])
        month = int(sys.argv[2])
        export_trades(year, month)
    else:
        # Default to current month
        export_trades()
